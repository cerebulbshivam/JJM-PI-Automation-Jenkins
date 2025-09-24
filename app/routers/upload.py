from fastapi.responses import JSONResponse
from fastapi import APIRouter, UploadFile, WebSocket
from fastapi.responses import HTMLResponse, StreamingResponse
import pandas as pd
import io
import time

from collections import defaultdict # Added this import
from app.database import get_db_connection
import pyodbc
from app.utils.utils import clean_value, derive_base_reservoir, normalize_reservoir, normalize_pressure_topic_id, normalize_other_topic_id, send_log
from app.utils.pi_tag_utility import create_pi_tag, create_pi_multiple_tag
from app.utils.mqtt_topics_utility import check_multiple_topics
from app.utils.topic_tags_json_utility import chlorine_tags_json_utility, fl_tags_json_utility, pressure_tags_json_utility
import json
import os
from app.utils.logger import setup_logger
# Get the root logger
logger = setup_logger()

router = APIRouter()

# --- Columns as provided from the SQL database schema ---
SQL_COLS = [
    "Site_name", "State", "Region", "Circle", "Division", "Sub_Division", "Block",
    "Scheme_ID", "Scheme_Name", "Water_Source", "Water_Supply_Well", "Type_of_Well",
    "Depth_of_Well", "Diameter_of_well", "Pump_House_Location_Geotag_Latitude",
    "Pump_House_Location_Geotag_Longitude", "Pump_House_Location_Geotag",
    "Numbers_of_Pumps", "Type_of_Source", "Availability_of_Existing_Energy_Meter",
    "Working_condition_of_Energy_Meter", "Scheme_Category", "Name_of_the_Incharge",
    "Position", "Scheme_Contact", "Village_Name", "Population", "Household",
    "Habitation", "Name_of_Base_Reservoir", "Name_of_the_Reservoir", "Reservoir_Type",
    "Number_of_ESR", "Number_of_MBR", "Number_of_GSR", "Reservoir_Capacity", "Operator_Name",
    "Reservoir_Contact", "Reservoir_Geotag_Latitude", "Reservoir_Geotag_Longitude",
    "Reservoir_Geotag_Info", "Name_of_Source", "Inlet_Line_Size", "Inlet_Line_Material",
    "Inlet_Rising_Main_Line_Size", "Inlet_Rising_Main_Line_Material", "Number_of_Outlet",
    "Outlet_Line_Size", "Outlet_Line_Material", "Outlet_Distribution_Line_Size",
    "Outlet_Distribution_Line_Material", "Availability_of_Isolation_Valve",
    "Geo_Location_of_Flow_Meter_Latitude", "Geo_Location_of_Flow_Meter_Longitude",
    "Geo_Location_of_Flow_Meter", "Geo_Location_of_RCA_Latitude",
    "Geo_Location_of_RCA_Longitude", "Geo_Location_of_RCA", "Number_of_Distribution",
    "Critical_Pressure_Point_Location_Latitude", "Critical_Pressure_Point_Location_Longitude",
    "Critical_Pressure_Point_Location", "Distribution_Line_Size",
    "Distribution_Line_Material", "Reservoir_Category", "Status", "Number_of_Reservoir",
    "Reservoir_Capacity_Val", "Reservoir_Level_Population",
    "Reservoir_Level_Household", "MBR_LPCD_RESERVOIR"
]

EXPECTED_MQTT_COLS = [
    "Region", "Circle", "Division", "Sub Division", "Block",
    "Schme ID  Name", "Village", "Reservoir"
]

# In-memory storage for files (for demonstration purposes)
temp_files = {}

logs = []  # global list to store logs

def log_message(msg: str):
    logs.append(msg)
    if len(logs) > 100:  # keep only recent 100 logs
        logs.pop(0)



# 1️⃣ Ingest Asset Metadata File into SQL (Step 1)
@router.post("/ingest_asset_metadata", response_class=HTMLResponse)
async def ingest_asset_metadata(file: UploadFile):
    conn = None
    try:
        start_time = time.time()
        
        df = pd.read_excel(file.file)
        logger.info(f"Read {len(df)} rows from the uploaded file.")
        #  if there is no data in df_verif
        if df.empty:
            logger.error("The uploaded file is empty.")
            return "<div class='text-red-600'>❌ The uploaded file is empty.</div>"

        # Step 1: Ensure all expected SQL_COLS exist 
        missing_cols_excel = [col for col in SQL_COLS if col not in df.columns]
        if missing_cols_excel:
            logger.error(f"Missing required columns: {', '.join(missing_cols_excel)}")
            return f"""
            <div class="text-red-600">Invalid Asset Metadata structure.</div>
            <p class="text-red-600">Missing required columns: {', '.join(missing_cols_excel)}</p>
            """

        # Remove duplicate rows in Excel based on unique key
        df = df.drop_duplicates(subset=["Name_of_the_Reservoir", "Village_Name"])
        file.file.seek(0)
        # Clear temporary files before starting
        
        
        conn = get_db_connection()
        cursor = conn.cursor()

        inserted, updated = 0, 0

        insert_cols_str = ", ".join([f"[{col}]" for col in SQL_COLS])
        insert_placeholders = ', '.join(['?'] * len(SQL_COLS))
        insert_sql = f"INSERT INTO dbo.Asset_Metadata_New ({insert_cols_str}) VALUES ({insert_placeholders})"

        update_sql = """
            UPDATE dbo.Asset_Metadata_New 
            SET [Status]=? 
            WHERE [Name_of_the_Reservoir]=? AND [Village_Name]=?
        """

        fetch_sql = """
            SELECT [Status]
            FROM dbo.Asset_Metadata_New
            WHERE [Name_of_the_Reservoir]=? AND [Village_Name]=?
        """

        for _, row in df.iterrows():
            # Prepare values for INSERT
            values = []
            for col in SQL_COLS:
                
                if col in df.columns:
                    if col == "Status":
                        values.append("Inactive")
                    else:
                        values.append(clean_value(row.get(col), col))
                else:
                    values.append(None)

            # Check if row exists
            cursor.execute(fetch_sql, row.get("Name_of_the_Reservoir"), row.get("Village_Name"))
            existing_row = cursor.fetchone()

            new_status = str(row.get("Status") or "").strip().lower()

            if existing_row:
                existing_status = (existing_row[0] or "").strip().lower()

                # Always update Status to 'Inactive'
                if existing_status != "Inactive":   
                    cursor.execute(
                        update_sql,
                        ("Inactive", row.get("Name_of_the_Reservoir"), row.get("Village_Name"))
                    )
                    updated += cursor.rowcount
            else:
                cursor.execute(insert_sql, *values)
                inserted += 1

        conn.commit()
        total_time = time.time() - start_time
        logger.info(f"New rows Inserted: {inserted} rows in {total_time:.2f} seconds.")
        # Inserted: {inserted}, Updated: {updated}
        return f"""
        <div class="bg-gray-900 text-green-400 font-mono p-4 rounded-md mb-4">
            Asset Metadata processed successfully.<br>
            New Insertions: {inserted}<br>
            Took {total_time:.2f} seconds.
        </div>
        """

    except Exception as e:
        if conn:
            conn.rollback()
        return f"<div class='text-red-600'> Error: {str(e)}</div>"
    finally:
        if conn:
            conn.close()

   
# 2️⃣ Validate Metadata File (Step 2)
@router.post("/validate_metadata", response_class=HTMLResponse)
async def validate_metadata(file: UploadFile):
    try:
        temp_files.clear()
        import io

        # Read verification file
        file_bytes = await file.read()
        filename = file.filename.lower()
        
        # Store the MQTT topics file content
        temp_files['mqtt_topics'] = file_bytes
        logger.info(f"{filename} is uploaded to validate data")
        # Always read the first sheet, since only one sheet will be present
        if filename.endswith(".xls"):
            df_verif = pd.read_excel(io.BytesIO(file_bytes), engine="xlrd")
        elif filename.endswith(".xlsx"):
            df_verif = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
        else:
            logger.error("Unsupported File format. Use .xls or .xlsx")
            return "<div class='text-red-600'>❌ Unsupported file format. Use .xls or .xlsx</div>"

        #  if there is no data in df_verif
        if df_verif.empty:
            logger.error("The uploaded file is empty.")
            return "<div class='text-red-600'>❌ The uploaded file is empty.</div>"

        # Connect to DB
        conn = get_db_connection()
        cursor = conn.cursor()

        
        df_verif['Validation_Status'] = 'Not Validated' # Initialize status column
        validation_results = []
        verified_count = 0

        for index, row in df_verif.iterrows():
            # Extract Scheme_ID and Scheme_Name
            scheme_id, scheme_name = None, None
            scheme_full = str(row['Schme ID  Name'])
            if "-" in scheme_full:
                scheme_id, scheme_name = scheme_full.split("-", 1)
            else:
                parts = scheme_full.split(maxsplit=1)
                scheme_id = parts[0]
                scheme_name = parts[1] if len(parts) > 1 else ""

            scheme_id = scheme_id.strip()
            scheme_name = scheme_name.strip()
            
            scheme_sub_div = str(row['Sub Division']).strip()
            if scheme_sub_div.upper().startswith("SUB DIVISION"):
                scheme_sub_div = scheme_sub_div.replace("Sub Division", "", 1).strip()

            # Normalize inputs
            scheme_name = scheme_name.upper()
            region = str(row["Region"]).strip().upper()
            circle = str(row["Circle"]).strip().upper()
            division = str(row["Division"]).strip().upper()
            scheme_sub_div = scheme_sub_div.upper()
            block = str(row["Block"]).strip().upper()
            village = str(row["Village"]).strip().upper()
            reservoir = normalize_reservoir(row["Reservoir"])

            # Update DB row to Active
            cursor.execute("""
                UPDATE dbo.Asset_Metadata_New
                SET Status = 'Validated'
                WHERE LTRIM(RTRIM(Scheme_ID)) = ?
                AND UPPER(LTRIM(RTRIM(Scheme_Name))) = ?
                AND UPPER(LTRIM(RTRIM(Region))) = ?
                AND UPPER(LTRIM(RTRIM(Circle))) = ?
                AND UPPER(LTRIM(RTRIM(Division))) = ?
                AND UPPER(LTRIM(RTRIM(Sub_Division))) = ?
                AND UPPER(LTRIM(RTRIM(Block))) = ?
                AND UPPER(LTRIM(RTRIM(Village_Name))) = ?
                AND UPPER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(Name_of_the_Reservoir)), '-', ''), ' ', ''), 'OL', 'OUTLET'), '', ''),'.','')) = ?
            """, (
                scheme_id,
                scheme_name,
                region,
                circle,
                division,
                scheme_sub_div,
                block,
                village,
                reservoir,
            ))

            if cursor.rowcount > 0:
                verified_count += cursor.rowcount
                df_verif.loc[index, 'Validation_Status'] = 'Validated'
            
        conn.commit()

        # Store the DataFrame with validation status in temp_files
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_verif.to_excel(writer, index=False, sheet_name='Validation_Status')
        output.seek(0)
        temp_files['validation_status_report'] = output.read()

        # Count Active & Inactive from DB
        
        active_count = verified_count  # Use verified_count for active rows
        inactive_count = len(df_verif) - verified_count  # Estimate inactive rows

        cursor.close()
        conn.close()
        logger.info(f"Validation completed. Active: {active_count}, Inactive: {inactive_count}")
        logger.info(f"Step 3 is triggered..!!")
        return f"""
        <div class="bg-gray-900 text-green-400 font-mono p-4 rounded-md mb-4 flex flex-col md:flex-row md:items-center md:justify-between">
            <div class="mb-2 md:mb-0">
                Verification & DB Updation Completed..!<br/>
                Validated Rows: {active_count} || Invalidated Rows: {inactive_count}<br/>
                Total Updated Rows: {active_count + inactive_count}
            </div>
            <div class="text-center">
                <a href="/download_validation_status" 
                class="inline-flex items-center px-4 py-1 border border-transparent text-sm font-medium rounded-md shadow-sm text-white bg-blue-600 hover:bg-blue-700 focus:outline-none focus:ring-2 focus:ring-offset-2 focus:ring-blue-500 transition duration-150 ease-in-out">
                    <svg xmlns="http://www.w3.org/2000/svg" class="h-5 w-5 mr-2" viewBox="0 0 20 20" fill="currentColor">
                        <path fill-rule="evenodd" d="M3 17a1 1 0 011-1h12a1 1 0 110 2H4a1 1 0 01-1-1zm3.293-7.707a1 1 0 011.414 0L9 10.586V3a1 1 0 112 0v7.586l1.293-1.293a1 1 0 111.414 1.414l-3 3a1 1 0 01-1.414 0l-3-3a1 1 0 010-1.414z" clip-rule="evenodd" />
                    </svg>
                    Data Validation Report
                </a>
            </div>
        </div>


        <script>
          document.getElementById('final-submit-step').removeAttribute('hidden');
          document.getElementById('submit-all-btn').disabled = false;
          document.getElementById('submit-all-btn').click();
        </script>
        """

    except Exception as e:
        logger.error(f"Error in validate_metadata: {str(e)}")
        return f"<div class='text-red-600'>❌ Error: Please Upload valid File..!!</div>"

# 3️⃣ create PI Tags Check topic in MQTT
@router.post("/final_upload", response_class=HTMLResponse)
async def final_upload():
    
    validation_file_content = temp_files.get('mqtt_topics')

    if not validation_file_content:
        return "<p class='text-red-600'>Error: First perform Validation at step 2</p>"
    
    try:
        
        try:
            # await send_log("📑 Reading topics file...")
            df_verif = pd.read_excel(io.BytesIO(validation_file_content), engine="openpyxl")
        except Exception:
            df_verif = pd.read_excel(io.BytesIO(validation_file_content), engine="xlrd")
        
        

        # await send_log("🔌 Connecting to database...")
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # await send_log("Fetching active data from database...")
        cursor.execute("""
            SELECT amn.[Site_name], amc.[region], amc.[Scheme_ID], amn.[Scheme_Name], amc.[village_name_cap], amc.[Reservoir_Cap], amn.[Village_Name], amn.[Name_of_the_Reservoir], amn.[Region], amn.[Name_of_Base_Reservoir]
            FROM [dbo].[Asset_Metadata_New] amn
            INNER JOIN [dbo].[Asset_Metadata_Count] amc
                ON amn.Scheme_ID = amc.Scheme_ID
            AND UPPER(LTRIM(RTRIM(amn.Village_Name))) = UPPER(LTRIM(RTRIM(amc.Village_Name)))
            AND UPPER(REPLACE(LTRIM(RTRIM(amn.Name_of_the_Reservoir)), ' ', '')) = 
                UPPER(REPLACE(LTRIM(RTRIM(amc.Name_of_the_Reservoir)), ' ', ''))
            WHERE amn.Status = 'Validated'
        """)
        rows = cursor.fetchall()
        rows = [tuple(r) for r in rows]

        active_df = pd.DataFrame(rows, columns=["Site_name","region","Scheme_ID","Scheme_Name" ,"village_name_cap", "Reservoir_Cap", "Village_Name", "Name_of_the_Reservoir", "Region","Name_of_Base_Reservoir"])

        if active_df.empty:
            return "<div class='text-red-600'>❌ No active records found in the database.</div>"
        
        tags = []
        multiple_tags_to_create = []
        all_pressure_tags_by_region = defaultdict(list)
        all_flow_meter_tags_by_region = defaultdict(list)
        all_chlorine_tags_by_region = defaultdict(list)
        
        tags_status_result = {'created': [], 'skipped': [], 'errors': []}
         # To store results of MQTT checks
        # tag_to_topic_map = {} # Map generated tags to their MQTT topics
        unique_topics_to_check = defaultdict(list) # Collect unique topics to avoid redundant checks

        # await send_log("=> Collecting Tags to create...")
        for _, r in active_df.iterrows():
            scheme_id = r["Scheme_ID"]
            current_region = r["Region"]
            village = r["village_name_cap"].upper().replace(" ", "_")
            reservoir = r["Reservoir_Cap"].upper().replace(" ", "_")
            
            test_village_name = r["Village_Name"].upper().strip()
            test_name_of_reservoir = normalize_reservoir(r["Name_of_the_Reservoir"]).upper()


            # Get topics from df_verif for this scheme_id (if exists)
            # filter df_verif
            topic_row = df_verif[
                (df_verif["Village"].str.upper().str.strip() == test_village_name) &
                (
                    (df_verif["Reservoir"].apply(normalize_reservoir).str.upper().str.strip() == test_name_of_reservoir) |
                    (df_verif["Reservoir"].str.upper().str.strip() == r["Name_of_Base_Reservoir"].upper().strip())
                )
            ]

            # ⛔ skip if not matching in df_verif
            if topic_row.empty:                
                logger.warning(f"No match in active_df for Village={test_village_name}, Reservoir={test_name_of_reservoir}")
                continue

            # get the first matching row
            topic_row = topic_row.head(1)
            topic_cl = str(topic_row["Topic For CL"].values[0]) if not topic_row.empty else ""
            cl_type = str(topic_row["CL Type"].values[0]) if not topic_row.empty else ""
            topic_flow = str(topic_row["Topic For Flow Meter"].values[0]) if not topic_row.empty else ""
            topic_press = str(topic_row["Topic For Pressure"].values[0]) if not topic_row.empty else ""

            cl_tag = f"JJM.MH_JJM_{scheme_id}_{village}_RES_{reservoir}_CL"
            fl_rate_tag =  f"JJM.MH_JJM_{scheme_id}_{village}_RES_{reservoir}_FL_RATE"
            total_fl_tag = f"JJM.MH_JJM_{scheme_id}_{village}_RES_{reservoir}_TOT_FL"
            press_tag = f"JJM.MH_JJM_{scheme_id}_{village}_RES_{reservoir}_PRESS"
            sen_err_cl_tag = f"JJM.MH_JJM_{scheme_id}_{village}_RES_{reservoir}_SEN_ERR_CL"
            sen_err_fl_mtr_tag = f"JJM.MH_JJM_{scheme_id}_{village}_RES_{reservoir}_SEN_ERR_FL_MTR"

            all_pressure_tags_by_region[current_region].append({"topic": topic_press, "tag": press_tag})
            all_flow_meter_tags_by_region[current_region].append({"topic": topic_flow, "fl_rate_tag": fl_rate_tag,"total_fl_tag" : total_fl_tag, "sen_err_fl_mtr_tag": sen_err_fl_mtr_tag})
            all_chlorine_tags_by_region[current_region].append({"topic": topic_cl,"cl_type" : cl_type , "cl_tag": cl_tag,"cl_error_tag" : sen_err_cl_tag})

            # Populate unique_topics_to_check and tag_to_topic_map
            if topic_flow:
                if topic_flow.endswith(".0"):
                    topic_flow = topic_flow[:-2]
                unique_topics_to_check["fm"].append(topic_flow)
            if topic_cl:
                if topic_cl.endswith(".0"):
                    topic_cl = topic_cl[:-2]
                unique_topics_to_check["cl"].append(topic_cl)
            if topic_press:
                if topic_press.endswith(".0"):
                    topic_press = topic_press[:-2]
                unique_topics_to_check["pressure"].append(topic_press)

            print(f"======================{current_region}========================")
            
            multiple_tags_to_create.extend([cl_tag, fl_rate_tag, total_fl_tag, press_tag, sen_err_cl_tag, sen_err_fl_mtr_tag])
        
            row = {
                "Scheme_ID": scheme_id,
                "village_name_cap": r["village_name_cap"],
                "Reservoir_Cap": r["Reservoir_Cap"],
                "Name_of_the_Reservoir": test_name_of_reservoir,
                "Village_Name": test_village_name,
                "CL_Tag": cl_tag,
                "FL_Rate_Tag": fl_rate_tag,
                "Tot_FL_Tag": total_fl_tag,
                "Press_Tag": press_tag,
                "Sen_Err_CL_Tag": sen_err_cl_tag,
                "Sen_Err_FL_MTR_Tag": sen_err_fl_mtr_tag,
                "Topic For Flow Meter" : topic_flow,
                "Topic For CL" : topic_cl,
                "Topic For Pressure" : topic_press,
            }
            result = await create_pi_multiple_tag([cl_tag, fl_rate_tag, total_fl_tag, press_tag, sen_err_cl_tag, sen_err_fl_mtr_tag], current_region)
           
                
            for key in ['created', 'skipped', 'errors']:
                for tag in result[key]:  # No need for .get() since we know the keys exist
                    if tag not in tags_status_result[key]:
                        tags_status_result[key].append(tag)

            tags.append(row)

        # --- Check MQTT data for unique topics ---
        # await send_log("Checking Topic IDs in MQTT...")
        mqtt_results = {}
        # logger.info(f"Checking MQTT data for {len(unique_topics_to_check)} unique topics...")
        # if len(unique_topics_to_check) > 0:
        #     mqtt_results = check_multiple_topics(unique_topics_to_check)
        # else:
        #     mqtt_results = {}
        if len(unique_topics_to_check) > 0:
            logger.info("==> Checking Topic IDs in MQTT...")
            mqtt_final_results = check_multiple_topics(unique_topics_to_check)
        else:
            mqtt_final_results = {}
        active_mqtt_checks = []
        mqtt_topic_results = {}
        # print(f"MQTT Results: {mqtt_results.keys()}")
        # Process MQTT results and update database records
    
        tags_df = pd.DataFrame(tags)
        for _, row in tags_df.iterrows():
            scheme_id = row['Scheme_ID']
            topics_to_check = [
                str(row['Topic For CL']).strip(),
                str(row['Topic For Flow Meter']).strip(),
                str(row['Topic For Pressure']).strip()
            ]
            
            # Check if any of the row's topics have MQTT data
            has_mqtt_data = False
            for topic in topics_to_check:
                if topic and topic in mqtt_final_results.keys():
                    result = mqtt_final_results[topic]
                    if result.get('status', 'not_communicated') == 'communicated':
                        has_mqtt_data = True
                        mqtt_topic_results[topic] = True
                        active_mqtt_checks.append(topic)
                        
                        break
                    else:
                        mqtt_topic_results[topic] = False
            # print(f"=====================MQTT Data Check for Scheme_ID={scheme_id}, Village={row['Village_Name']}, Reservoir={row['Name_of_the_Reservoir']}: {'Active' if has_mqtt_data else 'Inactive'}")
            if has_mqtt_data:
                # Update database to mark record as active
                village = str(row["Village_Name"]).strip().upper()
                reservoir = normalize_reservoir(row["Name_of_the_Reservoir"])
                cursor.execute("""
                    UPDATE dbo.Asset_Metadata_New
                    SET Status = 'Active'
                    WHERE Scheme_ID = ? AND UPPER(LTRIM(RTRIM(Village_Name))) = ? AND UPPER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(Name_of_the_Reservoir)), '-', ''), ' ', ''), 'OL', 'OUTLET'), '', ''),'.','')) = ?
                """, (scheme_id, village, reservoir))
            else:
                village = str(row["Village_Name"]).strip().upper()
                reservoir = normalize_reservoir(row["Name_of_the_Reservoir"])
                cursor.execute("""
                    UPDATE dbo.Asset_Metadata_New
                    SET Status = 'Inactive'
                    WHERE Scheme_ID = ? AND UPPER(LTRIM(RTRIM(Village_Name))) = ? AND UPPER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(Name_of_the_Reservoir)), '-', ''), ' ', ''), 'OL', 'OUTLET'), '', ''),'.','')) = ?
                """, (scheme_id, village, reservoir))
                
                
        conn.commit()
        # await send_log(f"✅ Found {len(active_mqtt_checks)} active MQTT topics")

        # --- Process tags by region after the loop ---
        no_of_tags = len(multiple_tags_to_create)
        
        
        track_flag = 0
        for region, pressure_tags_list in all_pressure_tags_by_region.items():
            pressure_tags_json_utility(df_verif, region, pressure_tags_list)
            track_flag += len(pressure_tags_list)
        # await send_log(f"✅ Updated pressure tags json ")

        track_flag = 0
        for region, flow_meter_tags_list in all_flow_meter_tags_by_region.items():
            fl_tags_json_utility(df_verif, region, flow_meter_tags_list)
            track_flag += len(flow_meter_tags_list)
        
        # await send_log(f"✅ Updated FL tags json ")
            
        track_flag = 0
        for region, chlorine_tags_list in all_chlorine_tags_by_region.items():
            chlorine_tags_json_utility(df_verif, region, chlorine_tags_list)
            track_flag += len(chlorine_tags_list)
        # await send_log(f"✅ Updated CL tags json ")
            
        print(f"cl_tags_result for {track_flag} : Processed")


        # Write Excel and apply cell-level color formatting (unchanged)
        tag_columns = [
            "CL_Tag",
            "FL_Rate_Tag",
            "Tot_FL_Tag",
            "Press_Tag",
            "Sen_Err_CL_Tag",
            "Sen_Err_FL_MTR_Tag",
            "Topic For Flow Meter",
            "Topic For CL",
            "Topic For Pressure"
        ]

        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            tags_df.to_excel(writer, index=False, sheet_name="Tags")
        output.seek(0)
        wb = load_workbook(output)
        ws = wb["Tags"]
        col_indices = {}
        for idx, cell in enumerate(ws[1], 1):
            if cell.value in tag_columns:
                col_indices[cell.value] = idx
        green_fill = PatternFill(start_color="6DEC5E", end_color="6DEC5E", fill_type="solid")
        grey_fill = PatternFill(start_color="9CA3AF", end_color="9CA3AF", fill_type="solid")
        red_fill = PatternFill(start_color="ED3822", end_color="ED3822", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_name, col_idx in col_indices.items():
                cell = row[col_idx - 1]
                tag_value = cell.value
                # If the column is a topic id column, check its MQTT status
                if col_name in ["Topic For Flow Meter", "Topic For CL", "Topic For Pressure"]:
                    if tag_value and str(tag_value) != 'nan':
                        pass
                        topic_value = str(tag_value)
                        if topic_value.endswith('.0'):
                            topic_value = topic_value[:-2]
                         
                        topic_status = mqtt_final_results[topic_value]['status']
                        if topic_status == 'communicated':
                            cell.fill = green_fill  # Topic has MQTT data
                        elif topic_status == 'not_communicated':
                            cell.fill = red_fill   # Topic checked but no data
                        elif topic_status == 'error':
                            cell.fill = yellow_fill    # Topic not found/invalid
                    else:
                        cell.fill = grey_fill  # No topic specified
                else:
                    # For PI tag columns
                    if tag_value in tags_status_result['created']:
                        cell.fill = green_fill      # New tag created
                    elif tag_value in tags_status_result['skipped']:
                        cell.fill = grey_fill       # Tag already existed
                    else:
                        cell.fill = red_fill        # Error creating tag
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        temp_files["active_meta_data"] = final_output.read()

        # await send_log("📝 Final report genrated...")
        logger.info("Final report genrated...")

        # <div class="bg-white p-6 rounded-lg shadow-md max-w-lg mx-auto">
        # --- END PRESSURE TOPIC TO TAG MAPPING ---
        return f"""
        <div class="bg-gray-900 text-green-400 font-mono p-4 rounded-md mb-4">
            <div class="space-y-2">
                <div class="flex items-center space-x-2 text-green-600 font-semibold">
                <svg xmlns="http://www.w3.org/2000/svg" class="h-6 w-6" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 12l2 2 4-4m6 2a9 9 0 11-18 0 9 9 0 0118 0z" />
                </svg>
                <span>Tags creation in PI SDK is Completed.</span>
                </div>

                <div class="flex items-center space-x-2 text-green-600 font-semibold">
                <svg xmlns="http://www.w3.org/2000/svg" class="h-6 w-6" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 12l2 2 4-4m6 2a9 9 0 11-18 0 9 9 0 0118 0z" />
                </svg>
                <span>MQTT Topic ID Checked & Status Updated in DB.</span>
                </div>
                
                <div class="flex items-center space-x-2 text-green-600 font-semibold">
                <svg xmlns="http://www.w3.org/2000/svg" class="h-6 w-6" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 12l2 2 4-4m6 2a9 9 0 11-18 0 9 9 0 0118 0z" />
                </svg>
                <span>Topic-Tags JSON is Updated.</span>
                </div>
            </div>

            <div class="mt-2 text-center">
                <a href="/download_active_tags" class="inline-flex items-center px-4 py-1 border border-transparent text-sm font-medium rounded-md shadow-sm text-white bg-blue-600 hover:bg-blue-700 focus:outline-none focus:ring-2 focus:ring-offset-2 focus:ring-blue-500 transition duration-150 ease-in-out">
                <svg xmlns="http://www.w3.org/2000/svg" class="h-5 w-5 mr-2" viewBox="0 0 20 20" fill="currentColor">
                    <path fill-rule="evenodd" d="M3 17a1 1 0 011-1h12a1 1 0 110 2H4a1 1 0 01-1-1zm3.293-7.707a1 1 0 011.414 0L9 10.586V3a1 1 0 112 0v7.586l1.293-1.293a1 1 0 111.414 1.414l-3 3a1 1 0 01-1.414 0l-3-3a1 1 0 010-1.414z" clip-rule="evenodd" />
                </svg>
                Tags Creation Status Report
                </a>
            </div>
            </div>
            
        """

    except Exception as e:
        error_message = f"❌ Error during final upload: {str(e)}"
        # await send_log(error_message)
        return f"""
        <div class="bg-black text-red-400 font-mono text-sm rounded-lg p-4 text-left shadow-md">
            {error_message}
        </div>
        """
    
@router.get("/download_updated_metadata")
async def download_updated_metadata():
    if "asset_metadata_updated" not in temp_files:
        return "<div class='text-red-600'>❌ No updated metadata available. Run validation first.</div>"
    
    return StreamingResponse(
        io.BytesIO(temp_files["asset_metadata_updated"]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=updated_metadata.xlsx"}
    )


## Removed get_excel_sheets endpoint: sheet selection is no longer needed
    
@router.get("/download_validation_status")
async def download_validation_status():
    if "validation_status_report" not in temp_files:
        return HTMLResponse("<div class='text-red-600'>❌ No validation status report available. Run validation first.</div>")
    
    excel_data = io.BytesIO(temp_files["validation_status_report"])
    return StreamingResponse(
        excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=validation_status_report.xlsx"}
    )
    
@router.get("/logs", response_class=HTMLResponse)
async def get_logs():
    return "<br>".join(logs) if logs else "No logs yet..."

@router.get("/download_active_tags")
async def download_active_tags():
    
    if "active_meta_data" not in temp_files:
        return HTMLResponse("<div class='text-red-600'>No active tags found.</div>")
    
    excel_data = io.BytesIO(temp_files["active_meta_data"])
    return StreamingResponse(
        excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Tags_creation_Status.xlsx"}
    )
    